import re
import pandas as pd
import sys
import os
from fuzzywuzzy import process, fuzz
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from num2words import num2words
from decimal import Decimal, ROUND_HALF_UP
import pandas as pd
import unicodedata

# FUNCIONES PARA GENERACIÓN DE PLANILLA

def cargar_archivo(ruta):
    extension = os.path.splitext(ruta)[-1].lower()
    if extension == ".csv":
        try:
            return pd.read_csv(ruta, sep=',')
        except Exception:
            return pd.read_csv(ruta)
    elif extension in [".xls", ".xlsx"]:
        return pd.read_excel(ruta)
    else:
        raise ValueError(f"Formato no soportado: {extension}")
    
def limpiar_docentes(df, col):
    df[col] = df[col].astype(str).str.strip()
    return df

def agrupar_y_calcular(df, datos_docentes, col_curso):
    agrupado = (
        df.groupby('docente')
          .agg(curso=(col_curso, lambda x: ' / '.join(x)),
               cantidad_cursos=(col_curso, 'count'))
          .reset_index()
    )
    nombres_base = datos_docentes['Docente'].tolist()
    agrupado['Docente'] = [
        process.extractOne(n, nombres_base, scorer=fuzz.token_sort_ratio)[0]
        if process.extractOne(n, nombres_base, scorer=fuzz.token_sort_ratio)[1] >= 85 else None
        for n in agrupado['docente']
    ]
    agrupado = agrupado.merge(
        datos_docentes[['Docente', 'Sede', 'Categoria (Letra)', 'Categoria (Monto)', 'N°. Ruc', 'Estado']],
        on='Docente', how='left'
    )
    agrupado['Curso Dictado'] = agrupado['Categoria (Monto)'] * agrupado['cantidad_cursos'] * 28
    agrupado['Diseño de Examenes'] = agrupado['Categoria (Monto)'] * agrupado['cantidad_cursos'] * 4
    return agrupado


def agregar_clasificacion(df, ruta_clasificacion, normalizar_texto):
    if os.path.exists(ruta_clasificacion):
        try:
            clasif_df = pd.read_excel(ruta_clasificacion, header=1)
            clasif_df['Docente'] = clasif_df['Docente'].astype(str).str.strip()
            clasif_df['docente_norm'] = clasif_df['Docente'].apply(normalizar_texto)
            df['docente_norm'] = df['Docente'].apply(normalizar_texto)
            df = df.merge(clasif_df[['docente_norm', 'Monto']], on='docente_norm', how='left')
            df['Examen Clasif.'] = df['Monto'].fillna(0)
            df.drop(columns=['docente_norm', 'Monto'], inplace=True)
        except Exception as e:
            print(f"⚠️ Error al leer archivo de clasificación: {e}")
            df['Examen Clasif.'] = 0
    else:
        df['Examen Clasif.'] = 0
    return df

def construir_tabla(df):
    tabla = pd.DataFrame({
        'N°': range(1, len(df) + 1),
        'Docente': df['Docente'],
        'Sede': df['Sede'],
        'Categoria (Letra)': df['Categoria (Letra)'],
        'Categoria (Monto)': df['Categoria (Monto)'],
        'N°. Ruc': df['N°. Ruc'],
        'Curso': df['curso'],
        'Curso Dictado': df['Curso Dictado'],
        'Extra Curso': 0,
        'Cantidad Cursos': df['cantidad_cursos'],
        'Diseño de Examenes': df['Diseño de Examenes'],
        'Examen Clasif.': df['Examen Clasif.'],
        'Sub Total Pago S/.': 0,
        'Estado': df['Estado']
    })
    tabla['Sub Total Pago S/.'] = (
        tabla['Curso Dictado'] + tabla['Extra Curso'] +
        tabla['Diseño de Examenes'] + tabla['Examen Clasif.']
    )
    return tabla

def ajustar_nivel(row):
        nivel = row['nivel']
        idioma = row['idioma']
        ciclo = row['ciclo']
        if nivel == 'General':
            if idioma == 'Inglés':
                return 'Posgrado Intermedio'
            elif idioma == 'Portugués':
                try:
                    ciclo_num = int(str(ciclo).strip())
                except Exception:
                    ciclo_num = None
                if ciclo_num is not None:
                    if 1 <= ciclo_num <= 4:
                        return 'Posgrado Básico'
                    elif 5 <= ciclo_num <= 8:
                        return 'Posgrado Intermedio'
        return nivel

def crear_df_carga(datos, estado_planilla):
    datos = datos.copy()

    datos['nivel'] = datos.apply(ajustar_nivel, axis=1)

    df = pd.DataFrame({
        'Dias': datos['dias'].apply(traducir_dias),
        'H. Inicio': datos['horainicio'].astype(str).str[:5],
        'H. Fin': datos['horafin'].astype(str).str[:5],
        'Idioma': datos['idioma'],
        'Nivel': datos['nivel'].str.replace('Ã¡', 'á', regex=False),
        'Ciclo': datos['ciclo'],
        'Curso': datos[['idioma', 'nivel', 'ciclo']].astype(str).agg(' '.join, axis=1),
        'Sede': datos['sede'],
        'Sec.': '',
        'Matr.': datos['matriculados'],
        'Docente': datos['docente'],
        'Modalidad': datos['modalidad'],
        'Estado Planilla': estado_planilla
    })
    df = df.sort_values(by='Docente').reset_index(drop=True)
    df.insert(0, 'N°', range(1, len(df) + 1))
    return df

def formatear_hoja(ws, titulo):
    ws.insert_rows(1)
    max_col = ws.max_column
    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    ws["A1"].value = titulo
    ws["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A1"].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    ws["A1"].font = Font(bold=True, color="ffffff", size=22)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def ordenar_hojas_excel(wb, hojas_ordenadas):
    hojas_existentes = wb.sheetnames
    nuevas_hojas = [hoja for hoja in hojas_ordenadas if hoja in hojas_existentes]
    for idx, hoja in enumerate(nuevas_hojas):
        wb._sheets.insert(idx, wb[hoja])
    wb._sheets = wb._sheets[:len(nuevas_hojas)] + [s for s in wb._sheets if s.title not in nuevas_hojas]
    return wb

# FUNCIONES PARA DOCUMENTOS DE FASE INICIAL Y FINAL

def normalizar_texto(texto):
    if pd.isna(texto):
        return ''
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize('NFKD', texto)
    return ''.join([c for c in texto if not unicodedata.combining(c)])


def limpiar_nombre_archivo(nombre):
    return re.sub(r'[\\/*?:"<>|]', "", nombre)


def limpiar_numero(valor):
    return "" if pd.isna(valor) else str(valor).split('.')[0]


def reemplazar_en_parrafos(documento, reemplazos):
    for parrafo in documento.paragraphs:
        for marcador, valor in reemplazos.items():
            if marcador in parrafo.text:
                texto_nuevo = parrafo.text.replace(marcador, valor)
                for run in parrafo.runs:
                    run.text = ''
                if parrafo.runs:
                    parrafo.runs[0].text = texto_nuevo

def reemplazar_en_tablas(documento, reemplazos):
    for tabla in documento.tables:
        for fila in tabla.rows:
            for celda in fila.cells:
                for parrafo in celda.paragraphs:
                    for marcador, valor in reemplazos.items():
                        if marcador in parrafo.text:
                            texto_nuevo = parrafo.text.replace(marcador, valor)
                            for run in parrafo.runs:
                                run.text = ''
                            if parrafo.runs:
                                parrafo.runs[0].text = texto_nuevo

def monto_a_letras(monto):
    try:
        monto = Decimal(str(monto)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        entero = int(monto)
        centavos = int((monto - Decimal(entero)) * 100)
        return f"{num2words(entero, lang='es')} y {centavos:02d}/100 soles"
    except Exception:
        return "N/A"

def redactar_cursos(cadena):
    if not isinstance(cadena, str):
        return "N/A"
    cursos = [c.strip() for c in cadena.split("/") if c.strip()]
    if not cursos:
        return "N/A"
    resultado = f"servicio de dictado de 28 horas de clases de {cursos[0]}"
    for curso in cursos[1:]:
        resultado += f", 28 horas de clases de {curso}"
    return resultado

def traducir_dias(dias_raw: str) -> str:
    dias_dict = {
        'MONDAY': 'Lun', 'TUESDAY': 'Mar', 'WEDNESDAY': 'Mié',
        'THURSDAY': 'Jue', 'FRIDAY': 'Vie', 'SATURDAY': 'Sáb', 'SUNDAY': 'Dom'
    }
    dias = dias_raw.strip('{}').split(',') if isinstance(dias_raw, str) else []
    return ', '.join(dias_dict.get(d.strip().upper(), d.strip()) for d in dias)

def ruta_absoluta_relativa(path_relativo):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, path_relativo)
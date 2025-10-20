from fuzzywuzzy import process
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from repositories import excel_repo

def _normalize_text(text):
    """Normaliza texto para comparaciones."""
    if not isinstance(text, str):
        return ""
    return " ".join(text.strip().upper().split())


def _find_header_row(worksheet, target_column):
    """Busca la fila que contiene los encabezados."""
    for r in range(1, min(worksheet.max_row, 15) + 1):
        values = [str(c.value).strip() if c.value is not None else "" for c in worksheet[r]]
        if any(v.upper() == target_column.upper() for v in values):
            return r
    return None


def _extract_control_names(worksheet, header_row):
    """Extrae nombres únicos del archivo de control."""
    nombres_control_unicos = set()
    
    # Encontrar columna de nombres
    nombre_col = None
    for col in range(1, worksheet.max_column + 1):
        header_val = str(worksheet.cell(header_row, col).value or "").strip().upper()
        if header_val == "APELLIDOS Y NOMBRES":
            nombre_col = col
            break
    
    if nombre_col:
        for row in range(header_row + 1, worksheet.max_row + 1):
            nombre_cell = worksheet.cell(row, nombre_col)
            if nombre_cell and nombre_cell.value:
                nombres_control_unicos.add(str(nombre_cell.value).strip())
    
    return nombres_control_unicos


def _is_zero(value, tolerance=1e-6):
    """Verifica si un valor es prácticamente cero."""
    return abs(float(value)) < tolerance

def actualizar_control_pagos(planilla_path, control_path, numero_armada):
    # Validar archivos usando repositorio
    if not excel_repo.exists(planilla_path):
        raise FileNotFoundError(f"No se encontró el archivo de planilla: {planilla_path}")
    if not excel_repo.exists(control_path):
        raise FileNotFoundError(f"No se encontró el archivo de control: {control_path}")

    # Leer planilla usando repositorio
    print("📊 Leyendo planilla...")
    df_planilla = excel_repo.read_sheet(planilla_path, sheet_name="Planilla_Generador")

    # Configuración de armadas
    ARMADAS_CONFIG = {
        "Primera": "Primera armada",
        "Segunda": "Segunda armada",
        "Tercera": "Tercera armada",
    }

    if numero_armada not in ARMADAS_CONFIG:
        raise ValueError(f"Número de armada inválido. Use: {list(ARMADAS_CONFIG.keys())}")
    col_armada_name = ARMADAS_CONFIG[numero_armada]

    # Validar columnas requeridas en planilla
    required_planilla_columns = ["Docente", "Total_pago"]
    missing_columns = [col for col in required_planilla_columns if col not in df_planilla.columns]
    if missing_columns:
        raise ValueError(f"Columnas faltantes en planilla: {missing_columns}")

    print("🔍 Preparando datos de planilla...")
    nombres_planilla = df_planilla["Docente"].dropna().tolist()

    def crear_mapeo_montos(df_planilla, nombres_planilla):
        """Crea mapeo entre nombres en control y montos en planilla usando fuzzy matching."""
        print("🔄 Creando mapeo de montos...")
        mapeo_montos = {}
        nombres_control_unicos = set()
        
        # Leer archivo de control para extraer nombres
        wb_temp = load_workbook(control_path, data_only=True)
        ws_temp = wb_temp.active
        
        # Buscar fila de encabezados
        header_row_temp = _find_header_row(ws_temp, "APELLIDOS Y NOMBRES")
        if not header_row_temp:
            raise RuntimeError("No se encontró la columna 'APELLIDOS Y NOMBRES' en el archivo de control")
        
        # Extraer nombres únicos del control
        nombres_control_unicos = _extract_control_names(ws_temp, header_row_temp)
        
        # Crear mapeo usando fuzzy matching
        for docente_control in nombres_control_unicos:
            if isinstance(docente_control, str):
                match = process.extractOne(docente_control, nombres_planilla)
                if match:
                    nombre_match, score = match
                    if score >= 85:  # Umbral de similitud
                        fila = df_planilla.loc[df_planilla["Docente"] == nombre_match, "Total_pago"]
                        mapeo_montos[docente_control] = float(fila.values[0]) if not fila.empty else None
                    else:
                        mapeo_montos[docente_control] = None
                else:
                    mapeo_montos[docente_control] = None
        
        print(f"✅ Mapeo creado: {len(mapeo_montos)} docentes procesados")
        return mapeo_montos

    mapeo_montos = crear_mapeo_montos(df_planilla, nombres_planilla)

    def monto_por_nombre(docente_control):
        """Obtiene el monto correspondiente a un docente."""
        if not isinstance(docente_control, str):
            return None
        return mapeo_montos.get(docente_control.strip(), None)

    # Procesar archivo de control
    print("📝 Procesando archivo de control...")
    wb = load_workbook(control_path)
    wb_values = load_workbook(control_path, data_only=True)
    ws = wb.active
    ws_values = wb_values.active

    # Encontrar fila de encabezados
    header_row = _find_header_row(ws, "APELLIDOS Y NOMBRES")
    if header_row is None:
        raise RuntimeError("No se encontró la fila de encabezados con 'APELLIDOS Y NOMBRES'.")

    headers = {}
    for cell in ws[header_row]:
        key = _normalize_text(cell.value)
        if key:
            headers[key] = cell.column

    required_columns = [
        "APELLIDOS Y NOMBRES",
        "MONTO TOTAL PARA CONTRATO S/",
        "PRIMERA ARMADA",
        "SEGUNDA ARMADA",
        "TERCERA ARMADA",
        col_armada_name.upper(),
        "SALDO RESTANTE",
    ]
    for titulo in required_columns:
        if _normalize_text(titulo) not in headers:
            raise RuntimeError(f"No se encontró la columna: {titulo}")

    # Mapear columnas
    column_mapping = {
        'nombre': headers["APELLIDOS Y NOMBRES"],
        'total': headers["MONTO TOTAL PARA CONTRATO S/"],
        'primera': headers["PRIMERA ARMADA"],
        'segunda': headers["SEGUNDA ARMADA"],
        'tercera': headers["TERCERA ARMADA"],
        'armada_actual': headers[col_armada_name.upper()],
        'saldo': headers["SALDO RESTANTE"]
    }

    # Letras de columnas para fórmulas
    col_letters = {
        'total': get_column_letter(column_mapping['total']),
        'primera': get_column_letter(column_mapping['primera']),
        'segunda': get_column_letter(column_mapping['segunda']),
        'tercera': get_column_letter(column_mapping['tercera'])
    }

    # Listas para mensajes de resultado
    resultados = {
        'terminados': [],
        'por_acabar': [],
        'excedentes': []
    }

    # Procesar filas de datos
    print("⚙️ Actualizando pagos...")
    first_data_row = header_row + 1
    
    for r in range(first_data_row, ws.max_row + 1):
        nombre = ws.cell(row=r, column=column_mapping['nombre']).value
        if not nombre or str(nombre).strip().upper() in {"APELLIDOS Y NOMBRES", ""}:
            continue

        monto = monto_por_nombre(str(nombre))
        if monto is None:
            continue  # Sin coincidencia en planilla

        # Leer valores actuales
        total = ws_values.cell(row=r, column=column_mapping['total']).value or 0
        primera = ws_values.cell(row=r, column=column_mapping['primera']).value or 0
        segunda = ws_values.cell(row=r, column=column_mapping['segunda']).value or 0
        tercera = ws_values.cell(row=r, column=column_mapping['tercera']).value or 0
    
        saldo_actual = float(total) - float(primera) - float(segunda) - float(tercera)

        # Asignar monto a la armada correspondiente
        if monto <= saldo_actual:
            ws.cell(row=r, column=column_mapping['armada_actual']).value = float(monto)
            excedente = 0
        else:
            ws.cell(row=r, column=column_mapping['armada_actual']).value = float(saldo_actual)
            excedente = monto - saldo_actual

        # Actualizar fórmula de saldo restante
        formula = f"={col_letters['total']}{r}-SUM({col_letters['primera']}{r},{col_letters['segunda']}{r},{col_letters['tercera']}{r})"
        ws.cell(row=r, column=column_mapping['saldo']).value = formula

        # Registrar excedente en columna nueva
        col_excedente = headers.get("EXCEDENTE")
        if not col_excedente:
            col_excedente = ws.max_column + 1
            ws.cell(row=header_row, column=col_excedente).value = "Excedente"
            headers["EXCEDENTE"] = col_excedente
        ws.cell(row=r, column=col_excedente).value = excedente

        # Recalcular saldo después de asignación
        nueva_primera = ws.cell(row=r, column=column_mapping['primera']).value or 0
        nueva_segunda = ws.cell(row=r, column=column_mapping['segunda']).value or 0
        nueva_tercera = ws.cell(row=r, column=column_mapping['tercera']).value or 0
        nuevo_saldo = float(total) - float(nueva_primera) - float(nueva_segunda) - float(nueva_tercera)

        # Clasificar resultados
        if _is_zero(nuevo_saldo) and _is_zero(excedente):
            resultados['terminados'].append(f"- {nombre}")

        if 0 < nuevo_saldo <= 1280:
            resultados['por_acabar'].append(f"- {nombre} - Saldo restante: {nuevo_saldo}")

        if excedente > 0:
            resultados['excedentes'].append(f"- {nombre} - Monto excedente: {excedente}")

    # Guardar archivo actualizado
    wb.save(control_path)
    print(f"✅ Archivo actualizado guardado en: {control_path}")

    # Mostrar resultados organizados
    _mostrar_resultados(resultados)


def _mostrar_resultados(resultados):
    """Muestra los resultados organizados por categorías."""
    if resultados['terminados']:
        print("\n=== CONTRATOS TERMINADOS ===")
        for mensaje in resultados['terminados']:
            print(mensaje)

    if resultados['por_acabar']:
        print("\n=== CONTRATOS POR ACABAR ===")
        for mensaje in resultados['por_acabar']:
            print(mensaje)

    if resultados['excedentes']:
        print("\n=== CONTRATOS CON EXCEDENTE ===")
        for mensaje in resultados['excedentes']:
            print(mensaje)

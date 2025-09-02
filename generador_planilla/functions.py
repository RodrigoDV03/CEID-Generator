import pandas as pd
import os
from fuzzywuzzy import process, fuzz
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import unicodedata

# FUNCIONES PARA GENERACIÓN DE PLANILLA

# Cache simple para archivos Excel
_excel_cache = {}

def limpiar_cache_excel():
    global _excel_cache
    _excel_cache.clear()

def cargar_excel_con_cache(ruta, sheet_name=0, header='infer'):
    cache_key = f"{ruta}_{sheet_name}_{header}"
    
    if cache_key not in _excel_cache:
        if header == 'infer':
            _excel_cache[cache_key] = pd.read_excel(ruta, sheet_name=sheet_name)
        else:
            _excel_cache[cache_key] = pd.read_excel(ruta, sheet_name=sheet_name, header=header)
    
    return _excel_cache[cache_key].copy()  # Retorna una copia para evitar modificaciones accidentales

def cargar_archivo(ruta):
    extension = os.path.splitext(ruta)[-1].lower()
    if extension == ".csv":
        try:
            return pd.read_csv(ruta, sep=',')
        except Exception:
            return pd.read_csv(ruta)
    elif extension in [".xls", ".xlsx"]:
        return pd.read_excel(ruta)
    else:
        raise ValueError(f"Formato no soportado: {extension}")
    
def limpiar_docentes(df, col):
    df[col] = df[col].astype(str).str.strip()
    return df

def aplicar_transformaciones_base(datos):
    datos_transformados = datos.copy()

    mask_general = datos_transformados['nivel'] == 'General'
    mask_ingles = datos_transformados['idioma'] == 'Inglés'
    mask_portugues = datos_transformados['idioma'] == 'Portugués'
    
    # Aplicar transformaciones vectorizadas donde sea posible
    datos_transformados['nivel_optimized'] = datos_transformados['nivel']
    
    # Caso específico para Inglés General
    datos_transformados.loc[mask_general & mask_ingles, 'nivel_optimized'] = 'Posgrado Intermedio'
    
    # Caso específico para Portugués General
    mask_portugues_general = mask_general & mask_portugues
    if mask_portugues_general.any():
        # Para Portugués, necesitamos evaluar los ciclos
        for idx in datos_transformados[mask_portugues_general].index:
            try:
                ciclo_num = int(str(datos_transformados.loc[idx, 'ciclo']).strip())
                if 1 <= ciclo_num <= 4:
                    datos_transformados.loc[idx, 'nivel_optimized'] = 'Posgrado Básico'
                elif 5 <= ciclo_num <= 8:
                    datos_transformados.loc[idx, 'nivel_optimized'] = 'Posgrado Intermedio'
            except:
                pass  # Mantener el valor original si hay error
    
    # Para casos restantes, aplicar la función original
    mask_restantes = ~(mask_general & (mask_ingles | mask_portugues))
    if mask_restantes.any():
        datos_transformados.loc[mask_restantes, 'nivel_optimized'] = datos_transformados.loc[mask_restantes].apply(ajustar_nivel, axis=1)
    
    datos_transformados['nivel'] = datos_transformados['nivel_optimized']
    datos_transformados.drop('nivel_optimized', axis=1, inplace=True)
    
    # Aplicar modalidad (mantener apply por ahora ya que la lógica es compleja)
    datos_transformados['modalidad'] = datos_transformados.apply(ajustar_modalidad, axis=1)
    
    # Construir curso de forma vectorizada
    datos_transformados['Curso'] = (
        datos_transformados['idioma'].astype(str) + ' ' + 
        datos_transformados['nivel'].astype(str) + ' ' + 
        datos_transformados['ciclo'].astype(str)
    )
    
    return datos_transformados

def crear_mapeo_fuzzy(nombres_input, nombres_base, umbral=85):
    mapeo = {}
    for nombre in nombres_input:
        if nombre not in mapeo:  # Evitar recálculos
            resultado = process.extractOne(nombre, nombres_base, scorer=fuzz.token_sort_ratio)
            mapeo[nombre] = resultado[0] if resultado[1] >= umbral else None
    return mapeo

def agrupar_y_calcular(df, datos_docentes, col_curso):
    agrupado = (df.groupby('docente').agg(curso=(col_curso, lambda x: ' / '.join(x)), cantidad_cursos=(col_curso, 'count')).reset_index())
    nombres_base = datos_docentes['Docente'].tolist()
    
    # Crear mapeo fuzzy una sola vez
    mapeo_docentes = crear_mapeo_fuzzy(agrupado['docente'].tolist(), nombres_base, umbral=85)
    
    # Aplicar mapeo usando el diccionario
    agrupado['Docente'] = agrupado['docente'].map(mapeo_docentes)
    
    agrupado = agrupado.merge(
        datos_docentes[['Docente', 'Sede', 'Categoria (Letra)', 'Categoria (Monto)', 'N°. Ruc', 'Estado']], on='Docente', how='left'
    )
    agrupado['Curso Dictado'] = agrupado['Categoria (Monto)'] * agrupado['cantidad_cursos'] * 28
    agrupado['Diseño de Examenes'] = agrupado['Categoria (Monto)'] * agrupado['cantidad_cursos'] * 4
    return agrupado


def agregar_clasificacion(df, ruta_clasificacion, normalizar_texto):
    if os.path.exists(ruta_clasificacion):
        try:
            # Usar cache para evitar lecturas múltiples
            clasif_df = cargar_excel_con_cache(ruta_clasificacion, sheet_name=0, header=1)
            clasif_df['Docente'] = clasif_df['Docente'].astype(str).str.strip()
            clasif_df['docente_norm'] = clasif_df['Docente'].apply(normalizar_texto)
            df['docente_norm'] = df['Docente'].apply(normalizar_texto)
            df = df.merge(clasif_df[['docente_norm', 'Monto']], on='docente_norm', how='left')
            df['Examen Clasif.'] = df['Monto'].fillna(0)
            df.drop(columns=['docente_norm', 'Monto'], inplace=True)
        except Exception as e:
            print(f"⚠️ Error al leer archivo de clasificación: {e}")
            df['Examen Clasif.'] = 0
    else:
        df['Examen Clasif.'] = 0
    return df

def construir_tabla(df):
    tabla = pd.DataFrame({
        'N°': range(1, len(df) + 1),
        'Docente': df['Docente'],
        'Sede': df['Sede'],
        'Categoria (Letra)': df['Categoria (Letra)'],
        'Categoria (Monto)': df['Categoria (Monto)'],
        'N°. Ruc': df['N°. Ruc'],
        'Curso': df['curso'],
        'Curso Dictado': df['Curso Dictado'],
        'Extra Curso': 0,
        'Cantidad Cursos': df['cantidad_cursos'],
        'Diseño de Examenes': df['Diseño de Examenes'],
        'Examen Clasif.': df['Examen Clasif.'],
        'Total Pago S/.': 0,
        'Estado': df['Estado']
    })
    tabla['Total Pago S/.'] = (
        tabla['Curso Dictado'] + tabla['Extra Curso'] +
        tabla['Diseño de Examenes'] + tabla['Examen Clasif.']
    )
    return tabla

def ajustar_nivel(row):
        nivel = row['nivel']
        idioma = row['idioma']
        ciclo = row['ciclo']
        if nivel == 'General':
            if idioma == 'Inglés':
                return 'Posgrado Intermedio'
            elif idioma == 'Portugués':
                try:
                    ciclo_num = int(str(ciclo).strip())
                except Exception:
                    ciclo_num = None
                if ciclo_num is not None:
                    if 1 <= ciclo_num <= 4:
                        return 'Posgrado Básico'
                    elif 5 <= ciclo_num <= 8:
                        return 'Posgrado Intermedio'
        return nivel

def ajustar_modalidad(row):
    dias = row['dias']
    hora_fin = row['horafin']
    modalidad = row['modalidad']

    if (hora_fin == '22:30:00'or hora_fin == '15:00:00') and (dias == '{MONDAY,TUESDAY,WEDNESDAY,THURSDAY}' or dias == '{SATURDAY,SUNDAY}'):
        modalidad = 'INTENSIVO VIRTUAL'

    return modalidad

def crear_df_carga(datos, estado_planilla):
    datos = datos.copy()

    if 'nivel' not in datos.columns:
        datos['nivel'] = datos.apply(ajustar_nivel, axis=1)
    if 'Curso' not in datos.columns:
        datos['Curso'] = datos[['idioma', 'nivel', 'ciclo']].astype(str).agg(' '.join, axis=1)

    df = pd.DataFrame({
        'Dias': datos['dias'].apply(traducir_dias),
        'H. Inicio': datos['horainicio'].astype(str).str[:5],
        'H. Fin': datos['horafin'].astype(str).str[:5],
        'Idioma': datos['idioma'],
        'Nivel': datos['nivel'].str.replace('Ã¡', 'á', regex=False),
        'Ciclo': datos['ciclo'],
        'Curso': datos['Curso'],
        'Sede': datos['sede'],
        'Sec.': '',
        'Matr.': datos['matriculados'],
        'Docente': datos['docente'],
        'Modalidad': datos['modalidad'],
        'Estado Planilla': estado_planilla
    })
    df = df.sort_values(by='Docente').reset_index(drop=True)
    df.insert(0, 'N°', range(1, len(df) + 1))
    return df

def ordenar_hojas_excel(wb, hojas_ordenadas):
    hojas_existentes = wb.sheetnames
    nuevas_hojas = [hoja for hoja in hojas_ordenadas if hoja in hojas_existentes]
    for idx, hoja in enumerate(nuevas_hojas):
        wb._sheets.insert(idx, wb[hoja])
    wb._sheets = wb._sheets[:len(nuevas_hojas)] + [s for s in wb._sheets if s.title not in nuevas_hojas]
    return wb

def traducir_dias(dias_raw: str) -> str:
    dias_dict = {
        'MONDAY': 'Lun', 'TUESDAY': 'Mar', 'WEDNESDAY': 'Mié',
        'THURSDAY': 'Jue', 'FRIDAY': 'Vie', 'SATURDAY': 'Sáb', 'SUNDAY': 'Dom'
    }
    dias = dias_raw.strip('{}').split(',') if isinstance(dias_raw, str) else []
    return ', '.join(dias_dict.get(d.strip().upper(), d.strip()) for d in dias)

def normalizar_texto(texto):
    if pd.isna(texto):
        return ''
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize('NFKD', texto)
    return ''.join([c for c in texto if not unicodedata.combining(c)])

# Cache para estilos de Excel para evitar recrearlos
_excel_styles_cache = {}

def get_excel_style(style_name):
    if style_name not in _excel_styles_cache:
        
        if style_name == "thin_border":
            _excel_styles_cache[style_name] = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        elif style_name == "header_fill":
            _excel_styles_cache[style_name] = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        elif style_name == "header_font":
            _excel_styles_cache[style_name] = Font(bold=True, color="ffffff", size=12)
        elif style_name == "title_font":
            _excel_styles_cache[style_name] = Font(bold=True, color="ffffff", size=22)
        elif style_name == "bold_font":
            _excel_styles_cache[style_name] = Font(bold=True)
        elif style_name == "center_alignment":
            _excel_styles_cache[style_name] = Alignment(horizontal='center', vertical='center')
        elif style_name == "title_alignment":
            _excel_styles_cache[style_name] = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif style_name == "money_format":
            _excel_styles_cache[style_name] = '"S/ "#,##0.00'
            
    return _excel_styles_cache[style_name]

def aplicar_formato_excel_optimizado(ws, max_col, titulo_fusionado, es_planilla=False):
    ws.insert_rows(1)
    rango_titulo = f"A1:{get_column_letter(max_col)}1"
    ws.merge_cells(rango_titulo)
    ws["A1"].value = titulo_fusionado
    ws["A1"].alignment = get_excel_style("title_alignment")
    ws["A1"].fill = get_excel_style("header_fill")
    ws["A1"].font = get_excel_style("title_font")

    thin_border = get_excel_style("thin_border")
    header_fill = get_excel_style("header_fill")
    header_font = get_excel_style("header_font")
    center_alignment = get_excel_style("center_alignment")
    bold_font = get_excel_style("bold_font")
    money_format = get_excel_style("money_format")

    header_cells = [ws.cell(row=2, column=col) for col in range(1, max_col + 1)]
    for celda in header_cells:
        celda.alignment = center_alignment
        celda.fill = header_fill
        celda.font = header_font

    max_row = ws.max_row
    for row_cells in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row_cells:
            cell.border = thin_border
            cell.alignment = center_alignment

    if es_planilla:
        fila_total = max_row + 1
        ws.merge_cells(f"A{fila_total}:G{fila_total}")
        celda_total = ws[f"A{fila_total}"]
        celda_total.alignment = center_alignment
        celda_total.fill = header_fill

        # Columnas a sumar - aplicar en lote
        columnas_sumar = ['H', 'I', 'J', 'K', 'L', 'M']
        
        for col in columnas_sumar:
            celda = ws[f"{col}{fila_total}"]
            celda.value = f"=SUM({col}3:{col}{fila_total-1})"
            celda.font = bold_font
            celda.alignment = center_alignment
            celda.border = thin_border

        columnas_moneda = ['E', 'H', 'K', 'L', 'M']
        
        for col in columnas_moneda:
            # Aplicar formato a todo el rango de una vez
            for row_num in range(3, fila_total + 1):
                ws[f"{col}{row_num}"].number_format = money_format

    ajustar_anchos_columnas_optimizado(ws, max_col)

def ajustar_anchos_columnas_optimizado(ws, max_col):
    anchos_predefinidos = {
        1: 5,   # N°
        2: 30,  # Docente
        3: 15,  # Sede
        4: 8,   # Categoria (Letra)
        5: 12,  # Categoria (Monto)
        6: 15,  # N°. Ruc
        7: 40,  # Curso
        8: 12,  # Curso Dictado
        9: 10,  # Extra Curso
        10: 12, # Cantidad Cursos
        11: 15, # Diseño de Examenes
        12: 12, # Examen Clasif.
        13: 15, # Total Pago S/.
        14: 10  # Estado
    }
    
    for col in range(1, min(max_col + 1, len(anchos_predefinidos) + 1)):
        column_letter = get_column_letter(col)
        if col in anchos_predefinidos:
            ws.column_dimensions[column_letter].width = anchos_predefinidos[col]
        else:
            # Ancho por defecto para columnas adicionales
            ws.column_dimensions[column_letter].width = 12

def procesar_formato_multiple_hojas(wb, hojas_con_titulo, numero_carga_letra, month):
    for hoja, titulo_fusionado in hojas_con_titulo:
        if hoja in wb.sheetnames:
            ws = wb[hoja]
            max_col = ws.max_column
            
            es_planilla = (hoja == f"{numero_carga_letra} Planilla {month}" or hoja == "Planilla consolidada")
            
            aplicar_formato_excel_optimizado(ws, max_col, titulo_fusionado, es_planilla)
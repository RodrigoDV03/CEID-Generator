from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Cache para estilos de Excel para evitar recrearlos
_excel_styles_cache = {}

def get_excel_style(style_name):
    if style_name not in _excel_styles_cache:
        
        if style_name == "thin_border":
            _excel_styles_cache[style_name] = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        elif style_name == "header_fill":
            _excel_styles_cache[style_name] = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        elif style_name == "header_font":
            _excel_styles_cache[style_name] = Font(bold=True, color="ffffff", size=12)
        elif style_name == "title_font":
            _excel_styles_cache[style_name] = Font(bold=True, color="ffffff", size=22)
        elif style_name == "bold_font":
            _excel_styles_cache[style_name] = Font(bold=True)
        elif style_name == "center_alignment":
            _excel_styles_cache[style_name] = Alignment(horizontal='center', vertical='center')
        elif style_name == "title_alignment":
            _excel_styles_cache[style_name] = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif style_name == "money_format":
            _excel_styles_cache[style_name] = '"S/ "#,##0.00'
            
    return _excel_styles_cache[style_name]

def aplicar_formato_excel_optimizado(ws, max_col, titulo_fusionado, es_planilla=False):
    ws.insert_rows(1)
    rango_titulo = f"A1:{get_column_letter(max_col)}1"
    ws.merge_cells(rango_titulo)
    ws["A1"].value = titulo_fusionado
    ws["A1"].alignment = get_excel_style("title_alignment")
    ws["A1"].fill = get_excel_style("header_fill")
    ws["A1"].font = get_excel_style("title_font")

    thin_border = get_excel_style("thin_border")
    header_fill = get_excel_style("header_fill")
    header_font = get_excel_style("header_font")
    center_alignment = get_excel_style("center_alignment")
    bold_font = get_excel_style("bold_font")
    money_format = get_excel_style("money_format")

    header_cells = [ws.cell(row=2, column=col) for col in range(1, max_col + 1)]
    for celda in header_cells:
        celda.alignment = center_alignment
        celda.fill = header_fill
        celda.font = header_font

    max_row = ws.max_row
    for row_cells in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row_cells:
            cell.border = thin_border
            cell.alignment = center_alignment

    if es_planilla:
        fila_total = max_row + 1
        ws.merge_cells(f"A{fila_total}:G{fila_total}")
        celda_total = ws[f"A{fila_total}"]
        celda_total.alignment = center_alignment
        celda_total.fill = header_fill

        # Detectar si la columna I es "Bono" para ajustar las columnas a sumar
        es_enero = False
        if ws.max_row >= 2:
            cell_value = ws.cell(row=2, column=9).value
            if cell_value and "Bono" in str(cell_value):
                es_enero = True
        
        # Columnas a sumar - ajustadas según si es enero o no
        if es_enero:
            # H=Curso Dictado, I=Bono, J=Extra Curso, K=Cantidad, L=Diseño, M=Examen, N=Servicio, O=Total
            columnas_sumar = ['H', 'I', 'J', 'L', 'M', 'N', 'O']  # No incluir K (Cantidad Cursos)
            columnas_moneda = ['E', 'H', 'I', 'L', 'M', 'N', 'O']
        else:
            # H=Curso Dictado, I=Extra Curso, J=Cantidad, K=Diseño, L=Examen, M=Servicio, N=Total
            columnas_sumar = ['H', 'I', 'K', 'L', 'M', 'N']  # No incluir J (Cantidad Cursos)
            columnas_moneda = ['E', 'H', 'K', 'L', 'M', 'N']
        
        for col in columnas_sumar:
            celda = ws[f"{col}{fila_total}"]
            celda.value = f"=SUM({col}3:{col}{fila_total-1})"
            celda.font = bold_font
            celda.alignment = center_alignment
            celda.border = thin_border

        for col in columnas_moneda:
            # Aplicar formato a todo el rango de una vez
            for row_num in range(3, fila_total + 1):
                ws[f"{col}{row_num}"].number_format = money_format

    ajustar_anchos_columnas_optimizado(ws, max_col)

def ajustar_anchos_columnas_optimizado(ws, max_col):
    # Detectar si la columna I es "Bono" (cuando es enero) o "Extra Curso" (resto del año)
    es_enero = False
    if ws.max_row >= 2:  # Verificar que haya headers
        cell_value = ws.cell(row=2, column=9).value
        if cell_value and "Bono" in str(cell_value):
            es_enero = True
    
    if es_enero:
        # Mapeo de columnas cuando hay Bono
        anchos_predefinidos = {
            1: 5,   # N°
            2: 30,  # Docente
            3: 15,  # Sede
            4: 8,   # Categoria (Letra)
            5: 12,  # Categoria (Monto)
            6: 15,  # N_Ruc
            7: 40,  # Curso
            8: 12,  # Curso Dictado
            9: 10,  # Bono
            10: 10, # Extra Curso
            11: 12, # Cantidad Cursos
            12: 15, # Diseño de Examenes
            13: 12, # Examen Clasif.
            14: 15, # Servicio Actualización
            15: 15, # Total Pago S/.
            16: 10  # Estado
        }
    else:
        # Mapeo de columnas sin Bono
        anchos_predefinidos = {
            1: 5,   # N°
            2: 30,  # Docente
            3: 15,  # Sede
            4: 8,   # Categoria (Letra)
            5: 12,  # Categoria (Monto)
            6: 15,  # N_Ruc
            7: 40,  # Curso
            8: 12,  # Curso Dictado
            9: 10,  # Extra Curso
            10: 12, # Cantidad Cursos
            11: 15, # Diseño de Examenes
            12: 12, # Examen Clasif.
            13: 15, # Servicio Actualización
            14: 15, # Total Pago S/.
            15: 10  # Estado
        }
    
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        if col in anchos_predefinidos:
            ws.column_dimensions[column_letter].width = anchos_predefinidos[col]
        else:
            # Ancho por defecto para columnas adicionales
            ws.column_dimensions[column_letter].width = 12

def procesar_formato_multiple_hojas(wb, titulo_hojas, numero_carga_letra, month):
    for hoja, titulo_fusionado in titulo_hojas:
        if hoja in wb.sheetnames:
            ws = wb[hoja]
            max_col = ws.max_column
            
            es_planilla = (hoja == f"{numero_carga_letra} Planilla {month}" or hoja == "Planilla consolidada")
            
            aplicar_formato_excel_optimizado(ws, max_col, titulo_fusionado, es_planilla)

def ordenar_hojas_excel(wb, hojas_ordenadas):
    hojas_existentes = wb.sheetnames
    nuevas_hojas = [hoja for hoja in hojas_ordenadas if hoja in hojas_existentes]
    for idx, hoja in enumerate(nuevas_hojas):
        wb._sheets.insert(idx, wb[hoja])
    wb._sheets = wb._sheets[:len(nuevas_hojas)] + [s for s in wb._sheets if s.title not in nuevas_hojas]
    return wb

def aplicar_formato_planilla_generador(wb, nombre_hoja="Planilla_Generador"):
    """
    Aplica formato especial a la hoja Planilla_Generador con colores alternados por docente
    para mejorar la legibilidad y facilitar la identificación visual.
    """
    if nombre_hoja not in wb.sheetnames:
        return
    
    ws = wb[nombre_hoja]
    
    # Colores alternados para docentes (tonos azul claro y gris claro)
    color_docente_1 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Azul muy claro
    color_docente_2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Gris muy claro
    
    # Colores por tipo de servicio (sutiles)
    colores_servicio = {
        'Curso Académico': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),      # Verde claro
        'Diseño de Exámenes': PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),   # Naranja claro
        'Examen de Clasificación': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # Púrpura claro
        'Servicio de Actualización': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # Rosa claro
    }
    
    # Borde más grueso para separar docentes
    borde_separador = Border(
        top=Side(style='medium', color='000000'),
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thin_border = get_excel_style("thin_border")
    bold_font = get_excel_style("bold_font")
    center_alignment = get_excel_style("center_alignment")
    
    # Buscar columna de Docente y Tipo_Servicio_Desc
    col_docente = None
    col_servicio = None
    
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col_idx).value
        if cell_value == 'Docente':
            col_docente = col_idx
        elif cell_value == 'Tipo_Servicio_Desc':
            col_servicio = col_idx
    
    if not col_docente:
        print("⚠️ No se encontró columna 'Docente' en Planilla_Generador")
        return
    
    # Aplicar formato por docente
    docente_actual = None
    color_actual = color_docente_1
    contador_docente = 0
    
    for row_idx in range(3, ws.max_row + 1):
        docente = ws.cell(row=row_idx, column=col_docente).value
        
        # Cambiar de color cuando cambia el docente
        if docente != docente_actual:
            docente_actual = docente
            contador_docente += 1
            color_actual = color_docente_1 if contador_docente % 2 == 1 else color_docente_2
            
            # Aplicar borde superior grueso para separar docentes
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).border = borde_separador
        
        # Aplicar color de fondo según tipo de servicio si está disponible
        tipo_servicio = ws.cell(row=row_idx, column=col_servicio).value if col_servicio else None
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Aplicar color según tipo de servicio (prioridad) o alternado por docente
            if tipo_servicio and tipo_servicio in colores_servicio:
                cell.fill = colores_servicio[tipo_servicio]
            else:
                cell.fill = color_actual
            
            # Mantener bordes
            if cell.border != borde_separador:
                cell.border = thin_border
            
            # Aplicar alineación centrada
            cell.alignment = center_alignment
    
    # Ajustar anchos de columna específicos para Planilla_Generador
    anchos_planilla_generador = {
        1: 5,   # N°
        2: 30,  # Docente
        3: 8,   # Servicio_Nro
        4: 20,  # Tipo_Servicio_Desc
        5: 40,  # Curso_Individual
        6: 15,  # Modalidad_Curso
        7: 12,  # Horas_Servicio
        8: 15,  # Monto_Individual
    }
    
    for col_idx, ancho in anchos_planilla_generador.items():
        if col_idx <= ws.max_column:
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = ancho
    
    # Resto de columnas con ancho estándar
    for col_idx in range(len(anchos_planilla_generador) + 1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 12
    
    print(f"✅ Formato especial aplicado a {nombre_hoja}")
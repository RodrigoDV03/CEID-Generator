import openpyxl
import pandas as pd
import os
import datetime
from fuzzywuzzy import process, fuzz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from .utils import *

def traducir_dias(dias_raw: str) -> str:
    dias_dict = {
        'MONDAY': 'Lun', 'TUESDAY': 'Mar', 'WEDNESDAY': 'Mié',
        'THURSDAY': 'Jue', 'FRIDAY': 'Vie', 'SATURDAY': 'Sáb', 'SUNDAY': 'Dom'
    }
    dias = dias_raw.strip('{}').split(',') if isinstance(dias_raw, str) else []
    return ', '.join(dias_dict.get(d.strip().upper(), d.strip()) for d in dias)

def generar_planilla(ruta_cursos: str, ruta_docentes: str, ruta_clasificacion: str, mes_seleccionado: str, numero_carga: int):

    try:
        extension = os.path.splitext(ruta_cursos)[-1].lower()
        if extension == ".csv":
            try:
                datos = pd.read_csv(ruta_cursos, sep=',')
            except Exception:
                datos = pd.read_csv(ruta_cursos)
        elif extension in [".xls", ".xlsx"]:
            datos = pd.read_excel(ruta_cursos)
        else:
            raise ValueError("Formato de cursos no soportado. Usa CSV o Excel.")
        
    
        datos_docentes = pd.read_excel(ruta_docentes, sheet_name="list")

        datos['docente'] = datos['docente'].astype(str).str.strip()
        datos_docentes['Docente'] = datos_docentes['Docente'].astype(str).str.strip()

        datos = datos[~datos['docente'].isin(['', ',', None])]
        datos['detalles_curso'] = datos[['idioma', 'nivel', 'ciclo']].astype(str).agg(' '.join, axis=1)

        agrupar = (
            datos.groupby('docente')
                 .agg(curso=('detalles_curso', lambda x: ' / '.join(x)),
                      cantidad_cursos=('detalles_curso', 'count'))
                 .reset_index()
        )

        nombres_base = datos_docentes['Docente'].tolist()
        agrupar['Docente'] = [
            process.extractOne(n, nombres_base, scorer=fuzz.token_sort_ratio)[0]
            if process.extractOne(n, nombres_base, scorer=fuzz.token_sort_ratio)[1] >= 85 else None
            for n in agrupar['docente']
        ]

        agrupar = agrupar.merge(
            datos_docentes[['Docente', 'Sede', 'Categoria (Letra)', 'Categoria (Monto)', 'N°. Ruc', 'Contrato o tercero']],
            on='Docente', how='left'
        )

        agrupar['Curso Dictado'] = agrupar['Categoria (Monto)'] * agrupar['cantidad_cursos'] * 28
        agrupar['Diseño de Examenes'] = agrupar['Categoria (Monto)'] * agrupar['cantidad_cursos'] * 4

        # Examen de clasificación (manejo opcional)
        if os.path.exists(ruta_clasificacion):
            try:
                clasif_df = pd.read_excel(ruta_clasificacion, header=1)
                clasif_df['Docente'] = clasif_df['Docente'].astype(str).str.strip()
                clasif_df['docente_norm'] = clasif_df['Docente'].apply(normalizar_texto)

                agrupar['docente_norm'] = agrupar['Docente'].apply(normalizar_texto)
                agrupar = agrupar.merge(clasif_df[['docente_norm', 'Monto']], on='docente_norm', how='left')
                agrupar['Examen Clasif.'] = agrupar['Monto'].fillna(0)
                agrupar.drop(columns=['docente_norm', 'Monto'], inplace=True)
            except Exception as e:
                print(f"⚠️ Error al leer archivo de clasificación: {e}")
                agrupar['Examen Clasif.'] = 0
        else:
            print("⚠️ No se encontró el archivo de clasificación. Se asumirá monto 0 para todos.")
            agrupar['Examen Clasif.'] = 0

        TABLA = pd.DataFrame({
            'N°': range(1, len(agrupar) + 1),
            'Docente': agrupar['Docente'],
            'Sede': agrupar['Sede'],
            'Categoria (Letra)': agrupar['Categoria (Letra)'],
            'Categoria (Monto)': agrupar['Categoria (Monto)'],
            'N°. Ruc': agrupar['N°. Ruc'],
            'Curso': agrupar['curso'],
            'Curso Dictado': agrupar['Curso Dictado'],
            'Extra Curso': 0,
            'Cantidad Cursos': agrupar['cantidad_cursos'],
            'Diseño de Examenes': agrupar['Diseño de Examenes'],
            'Examen Clasif.': agrupar['Examen Clasif.'],
            'Sub Total Pago S/.': 0,
            'Contrato o Tercero': agrupar['Contrato o tercero']
        })
        TABLA['Sub Total Pago S/.'] = (
            TABLA['Curso Dictado'] + TABLA['Extra Curso'] +
            TABLA['Diseño de Examenes'] + TABLA['Examen Clasif.']
        )

        año_actual = datetime.datetime.now().year

        estado_planilla = "Primera planilla" if numero_carga == 1 else "Segunda planilla"
        nombre_hoja_carga = f"{numero_carga} carga académica"

        carpeta_salida = f"{mes_seleccionado} {año_actual}"
        os.makedirs(carpeta_salida, exist_ok=True)
        nombre_salida = f"{estado_planilla} {mes_seleccionado} {año_actual}.xlsx"
        ruta_salida = os.path.join(carpeta_salida, nombre_salida)

        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            TABLA.to_excel(writer, sheet_name=f"Planilla {mes_seleccionado}", index=False)
            columnas_extra = ['Idioma', 'Nro. Documento', 'Celular', 'Dirección', 'Correo personal', 'N° Contrato']
            hoja_generador = (
                TABLA
                .merge(datos_docentes[['Docente'] + columnas_extra], on='Docente', how='left')
                .rename(columns={
                    'Cantidad Cursos': 'Cantidad_cursos',
                    'Examen Clasif.': 'Examen_clasif',
                    'Categoria (Letra)': 'Categoria_letra',
                    'Categoria (Monto)': 'Categoria_monto',
                    'Sub Total Pago S/.': 'Subtotal_pago',
                    'Contrato o Tercero': 'Contrato_o_tercero',
                    'Idioma': 'Docente_idioma',
                    'N°. Ruc': 'N_Ruc',
                    'Nro. Documento': 'Numero_dni',
                    'Celular': 'Numero_celular',
                    'Dirección': 'Domicilio_docente',
                    'Correo personal': 'Correo_personal',
                    'N° Contrato': 'Nro_contrato'
                })
            )
            hoja_generador['Numero_dni'] = hoja_generador['Numero_dni'].apply(lambda x: str(int(float(x))).zfill(8) if pd.notna(x) else '')
            hoja_generador.to_excel(writer, sheet_name="Planilla_Generador", index=False)

            df_carga = pd.DataFrame({
                'Dias': datos['dias'].apply(traducir_dias),
                'H. Inicio': datos['horainicio'].astype(str).str[:5],
                'H. Fin': datos['horafin'].astype(str).str[:5],
                'Idioma': datos['idioma'],
                'Nivel': datos['nivel'].str.replace('Ã¡', 'á', regex=False),
                'Ciclo': datos['ciclo'],
                'Curso': datos[['idioma', 'nivel', 'ciclo']].astype(str).agg(' '.join, axis=1),
                'Sede': datos['sede'],
                'Sec.': '',
                'Matr.': datos['matriculados'],
                'Docente': datos['docente'],
                'Modalidad': datos['modalidad'],
                'Estado Planilla': estado_planilla
            })           
            if os.path.exists(ruta_clasificacion):
                try:
                    hoja_clasificacion = pd.read_excel(ruta_clasificacion, header=1)
                    hoja_clasificacion.to_excel(writer, sheet_name="Examen de clasificación", index=False)
                except Exception as e:
                    print(f"⚠️ Error al copiar hoja de clasificación: {e}")
            df_carga = df_carga.sort_values(by='Docente', ascending=True).reset_index(drop=True)
            df_carga.insert(0, 'N°', range(1, len(df_carga) + 1))
            df_carga.to_excel(writer, sheet_name=nombre_hoja_carga, index=False)

        wb = load_workbook(ruta_salida)
        titulo_fusionado = (
            "CENTRO DE IDIOMAS - FLCH - UNMSM\n"
            f"{numero_carga} CARGA ACADÉMICA - PERIODO {mes_seleccionado.upper()} {año_actual}\n"
            "MODALIDAD: VIRTUAL Y PRESENCIAL"
        )

        hojas_con_titulo = [
            ("Examen de clasificación", 
            f"CENTRO DE IDIOMAS - FLCH - UNMSM\nEXAMEN DE CLASIFICACIÓN - PERIODO {mes_seleccionado.upper()} {año_actual}\nMODALIDAD: VIRTUAL Y PRESENCIAL"),
            
            (f"Planilla {mes_seleccionado}", 
            f"CENTRO DE IDIOMAS - FLCH - UNMSM\n{numero_carga} PLANILLA - PERIODO {mes_seleccionado.upper()} {año_actual}\nMODALIDAD: VIRTUAL Y PRESENCIAL"),
            
            (nombre_hoja_carga, 
            f"CENTRO DE IDIOMAS - FLCH - UNMSM\n{numero_carga} CARGA ACADÉMICA - PERIODO {mes_seleccionado.upper()} {año_actual}\nMODALIDAD: VIRTUAL Y PRESENCIAL")
        ]

        for hoja, titulo_fusionado in hojas_con_titulo:
            if hoja in wb.sheetnames:
                ws = wb[hoja]
                ws.insert_rows(1)
                max_col = ws.max_column
                rango = f"A1:{get_column_letter(max_col)}1"
                ws.merge_cells(rango)
                ws["A1"].value = titulo_fusionado
                ws["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws["A1"].fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                ws["A1"].font = Font(bold=True, color="ffffff", size=22)

                # Aplicar bordes, centrado y ajuste de columnas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Ajustar ancho de columnas automáticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            try:
                                cell_len = len(str(cell.value))
                                if cell_len > max_length:
                                    max_length = cell_len
                            except:
                                pass
                    adjusted_width = max_length + 2  # Un poco de margen extra
                    ws.column_dimensions[col_letter].width = adjusted_width

            for col in range(1, max_col + 1):
                celda = ws.cell(row=2, column=col)
                celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                celda.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                celda.font = Font(bold=True, color="ffffff", size=12)

            if hoja == f"Planilla {mes_seleccionado}":
                fila_total = ws.max_row + 1

                # Escribir la palabra "Totales" en la columna A y combinar de A hasta G
                ws.merge_cells(f"A{fila_total}:G{fila_total}")
                celda_total = ws[f"A{fila_total}"]
                celda_total.alignment = Alignment(horizontal="center", vertical="center")
                celda_total.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

                # Insertar fórmulas para H a M
                columnas_sumar = ['H', 'I', 'J', 'K', 'L', 'M']
                for col in columnas_sumar:
                    celda = ws[f"{col}{fila_total}"]
                    celda.value = f"=SUM({col}3:{col}{fila_total-1})"
                    celda.font = Font(bold=True)
                    celda.alignment = Alignment(horizontal="center", vertical="center")

        hojas_ordenadas = ["Examen de clasificación", nombre_hoja_carga, f"Planilla {mes_seleccionado}", "Planilla_Generador"]
        hojas_existentes = wb.sheetnames
        nuevas_hojas = [hoja for hoja in hojas_ordenadas if hoja in hojas_existentes]
        for idx, hoja in enumerate(nuevas_hojas):
            wb._sheets.insert(idx, wb[hoja])
        wb._sheets = wb._sheets[:len(nuevas_hojas)] + [s for s in wb._sheets if s.title not in nuevas_hojas]
        
        wb.save(ruta_salida)
        return f"✅ {nombre_salida} generado correctamente."

    except Exception as e:
        return f"❌ Error: {e}"